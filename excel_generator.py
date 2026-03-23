"""
Excel Generator module for GMRIT Results
Creates formatted Excel files with student data
"""
import os
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from config import EXCEL_FILE, HTML_DIR
from logger import logger

class ExcelGenerator:
    """Generate Excel files from scraped data"""
    
    # Color scheme
    HEADER_BG = "4472C4"  # Blue
    HEADER_FONT = "FFFFFF"  # White
    ALT_ROW_BG = "D9E2F3"  # Light blue
    BORDER_COLOR = "000000"
    
    def __init__(self):
        self.workbook = None
        
    def grade_to_points(self, grade):
        """Convert grade to points for SGPA calculation"""
        mapping = {
            'S': 10, 'A': 9, 'B': 8, 'C': 7, 'D': 6, 'E': 5, 'F': 0,
        }
        return mapping.get(grade.strip(), 0)
    
    def extract_data_from_html(self, html_path):
        """Extract results data from HTML file"""
        if not os.path.exists(html_path):
            return None
            
        with open(html_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        return self._parse_html(html_content)
    
    def _parse_html(self, html_content):
        """Parse HTML content and extract student data"""
        soup = BeautifulSoup(html_content, 'html.parser')
        results = {}
        
        # Extract CGPA
        cgpa_div = soup.find('div', class_='MuiBox-root css-bmlw8o')
        if cgpa_div:
            cgpa_text = cgpa_div.get_text(strip=True)
            if 'CGPA' in cgpa_text:
                cgpa_value = cgpa_text.split('CGPA :')[-1].strip()
                results['CGPA'] = cgpa_value
        
        # Extract Student Name
        student_name_divs = soup.find_all('div', class_='MuiStack-root css-1yeei81')
        for div in student_name_divs:
            h6 = div.find('h6', string='Student Name')
            if h6:
                p = h6.find_next_sibling('p')
                if p:
                    results['Student Name'] = p.get_text(strip=True)
                    break
        
        # Find all semester sections
        semester_sections = soup.find_all('div', class_='MuiStack-root css-190s4gn')
        for section in semester_sections:
            sem_heading = section.find('h6', class_='MuiTypography-root MuiTypography-h6 css-1vblau3')
            if sem_heading:
                sem_title = sem_heading.get_text(strip=True)
                sem_key = sem_title.replace('Semester - ', 'Semester ')
                results[sem_key] = {'rows': [], 'SGPA': 'N/A'}
                
                # Find table
                table = section.find('table')
                if table:
                    rows = table.find_all('tr')
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) >= 6:
                            row_data = [cell.get_text(strip=True) for cell in cells]
                            if row_data[0].isdigit():
                                results[sem_key]['rows'].append(row_data)
                
                # Extract SGPA
                sgpa_div = section.find('div', class_='MuiBox-root css-eb1jj5')
                if sgpa_div:
                    sgpa_text = sgpa_div.get_text(strip=True)
                    if 'SGPA' in sgpa_text:
                        sgpa_value = sgpa_text.split('SGPA :')[-1].strip()
                        try:
                            results[sem_key]['SGPA'] = float(sgpa_value)
                        except ValueError:
                            results[sem_key]['SGPA'] = sgpa_value
        
        return results
    
    def process_all_html_files(self):
        """Process all HTML files in the output directory"""
        excel_data = {}
        
        # Get all HTML files (both _page_source.html and _page_source_after.html)
        html_files = [f for f in os.listdir(HTML_DIR) if f.endswith('.html')]
        
        processed_halltickets = set()
        
        for html_file in html_files:
            # Extract hallticket from filename (format: HALLTICKET_page_source_after.html)
            parts = html_file.split('_')
            if len(parts) >= 1:
                ht = parts[0]
                
                # Skip if already processed
                if ht in processed_halltickets:
                    continue
                
                # Prefer _after.html files (contain results), otherwise use _page_source.html
                if 'after' in html_file or not any(f.startswith(ht) and 'after' in f for f in html_files):
                    html_path = os.path.join(HTML_DIR, html_file)
                    
                    logger.info(f"Extracting data from {ht}...")
                    data = self.extract_data_from_html(html_path)
                    
                    if data:
                        excel_data[ht] = data
                        processed_halltickets.add(ht)
                        logger.success(f"✓ {ht}: {data.get('Student Name', 'N/A')} - CGPA: {data.get('CGPA', 'N/A')}")
        
        return excel_data
    
    def generate_excel(self, excel_data):
        """Generate Excel file from extracted data"""
        if not excel_data:
            logger.warning("No data to save to Excel")
            return False
        
        self.workbook = openpyxl.Workbook()
        
        # Collect all semesters and subjects
        all_semesters = set()
        all_subjects = set()
        
        for ht, data in excel_data.items():
            for sem, sem_data in data.items():
                if sem not in ['CGPA', 'Student Name']:
                    all_semesters.add(sem)
                    for row in sem_data.get('rows', []):
                        if len(row) >= 6:
                            all_subjects.add(row[1])
        
        all_semesters = sorted(all_semesters, key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 0)
        all_subjects = sorted(all_subjects)
        
        # Create CGPA sheet
        self._create_cgpa_sheet(excel_data, all_semesters)
        
        # Create subject sheets
        self._create_subject_sheets(excel_data, all_subjects)
        
        # Create Semester Details sheet
        self._create_details_sheet(excel_data)
        
        # Save workbook
        self.workbook.save(EXCEL_FILE)
        logger.success(f"✓ Excel file saved: {EXCEL_FILE}")
        return True
    
    def _create_cgpa_sheet(self, excel_data, all_semesters):
        """Create CGPA summary sheet"""
        ws = self.workbook.active
        ws.title = "CGPA"
        
        # Headers
        headers = ["Hallticket", "Student Name", "CGPA"] + [f"{sem} SGPA" for sem in all_semesters]
        ws.append(headers)
        
        # Data
        for ht, data in sorted(excel_data.items()):
            student_name = data.get('Student Name', 'N/A')
            cgpa = data.get('CGPA', 'N/A')
            row = [ht, student_name, cgpa]
            for sem in all_semesters:
                sgpa = data.get(sem, {}).get('SGPA', '')
                row.append(sgpa)
            ws.append(row)
        
        # Apply styling
        self._style_worksheet(ws, headers)
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        for i in range(4, len(headers) + 1):
            ws.column_dimensions[chr(64 + i)].width = 12
    
    def _create_subject_sheets(self, excel_data, all_subjects):
        """Create individual subject sheets"""
        subject_sheets = {}
        
        for subject_code in all_subjects:
            ws = self.workbook.create_sheet(subject_code)
            ws.append(["Hallticket", "Semester", "Subject Name", "Grade", "Credits", "Points"])
            subject_sheets[subject_code] = ws
        
        # Populate data
        for ht, data in excel_data.items():
            for sem, sem_data in data.items():
                if sem not in ['CGPA', 'Student Name']:
                    for row in sem_data.get('rows', []):
                        if len(row) >= 6:
                            subject_code = row[1]
                            subject_name = row[2]
                            grade = row[4]
                            credits = row[3]
                            points = self.grade_to_points(grade)
                            subject_sheets[subject_code].append([ht, sem, subject_name, grade, credits, points])
        
        # Style each subject sheet
        for ws in subject_sheets.values():
            headers = ["Hallticket", "Semester", "Subject Name", "Grade", "Credits", "Points"]
            self._style_worksheet(ws, headers)
            
            # Column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 8
    
    def _create_details_sheet(self, excel_data):
        """Create Semester Details sheet"""
        ws = self.workbook.create_sheet("Semester Details")
        ws.append(["Hallticket", "Semester", "Subject Code", "Subject Name", "Grade", "Credits", "Points"])
        
        all_details = []
        for ht, data in excel_data.items():
            for sem, sem_data in data.items():
                if sem not in ['CGPA', 'Student Name']:
                    for row in sem_data.get('rows', []):
                        if len(row) >= 6:
                            all_details.append([
                                ht, sem, row[1], row[2], row[4], row[3],
                                self.grade_to_points(row[4])
                            ])
        
        # Sort by subject code
        all_details.sort(key=lambda x: x[2])
        
        for row in all_details:
            ws.append(row)
        
        # Style
        headers = ["Hallticket", "Semester", "Subject Code", "Subject Name", "Grade", "Credits", "Points"]
        self._style_worksheet(ws, headers)
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 8
    
    def _style_worksheet(self, ws, headers):
        """Apply styling to a worksheet"""
        bold_font = Font(bold=True, color=self.HEADER_FONT)
        thin_border = Border(
            left=Side(style='thin', color=self.BORDER_COLOR),
            right=Side(style='thin', color=self.BORDER_COLOR),
            top=Side(style='thin', color=self.BORDER_COLOR),
            bottom=Side(style='thin', color=self.BORDER_COLOR)
        )
        header_fill = PatternFill(start_color=self.HEADER_BG, end_color=self.HEADER_BG, fill_type="solid")
        alt_fill = PatternFill(start_color=self.ALT_ROW_BG, end_color=self.ALT_ROW_BG, fill_type="solid")
        
        # Style header row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style data rows
        for row in range(2, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                # Name column left-aligned, others centered
                cell.alignment = Alignment(horizontal="left" if col == 2 else "center", vertical="center")
                if row % 2 == 0:
                    cell.fill = alt_fill
